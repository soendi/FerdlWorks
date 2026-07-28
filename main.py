import sys
import os
import math
import subprocess
import customtkinter as ctk
import tkinter as tk
from tkinter import messagebox, ttk
from PIL import Image
from datetime import datetime, timedelta

from lib.logger import setup_logger, get_logger, get_log_path
from lib.registry import reg_write, reg_read, reg_delete_all
from lib.icon import create_icon
from lib.database import get_db
from lib.settings_dialog import SettingsDialog
from lib.customer_dialog import CustomerDialog
from lib.customer_database import CustomerDatabase
from lib.tool_database import ToolDatabase
from lib.material_database import MaterialDatabase
from lib.text_database import TextDatabase
from lib.pdf_gen import generate_pdf
from lib.email_sender import send_email
from lib.updater import check_for_update, download_installer, install_update, install_and_restart
from lib.autostart import autostart_enable, autostart_disable, autostart_is_enabled
from lib.cloud_backup import gdrive_backup, gdrive_authorize, onedrive_backup, onedrive_authorize
from version import VERSION, APP_NAME, COMPANY_NAME
from lib.icon import set_window_icon

THEME_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "assets", "ferdlworks_theme.json")


class PositionItem:
    def __init__(self, pos_type, ref_id, description, quantity, unit, price_per_unit, total, extra_data=None):
        self.pos_type = pos_type
        self.ref_id = ref_id
        self.description = description
        self.quantity = quantity
        self.unit = unit
        self.price_per_unit = price_per_unit
        self.total = total
        self.extra_data = extra_data or {}


class FerdlWorksApp(ctk.CTk):
    def __init__(self, master_mode=False):
        super().__init__()
        self.logger = get_logger()
        self.db = get_db()
        self._master_mode = master_mode
        self.title(f"{APP_NAME} v{VERSION}")
        self._icon_path = create_icon()
        self.minsize(800, 620)
        self.geometry("1024x900")
        self._current_doc_id = None
        self._positions = []
        self._editing_pos_idx = None
        self._nav_arrow_art = False
        self._nav_arrow_cust = False
        self._build_menu()
        self._build_ui()
        self.bind("<Button-1>", self._on_global_click, add="+")
        self._new_doc()
        self.after(100, lambda: self._set_icon())
        self.after(500, self._check_overdue)
        self.logger.info(f"{APP_NAME} v{VERSION} gestartet (Master-Mode: {master_mode})")
        self.protocol("WM_DELETE_WINDOW", self._on_close)

    def _set_icon(self):
        paths = [self._icon_path]
        if getattr(sys, "frozen", False):
            base = os.path.dirname(sys.executable)
            paths += [os.path.join(base, "ferdlworks.ico"),
                      os.path.join(base, "_internal", "ferdlworks.ico"),
                      os.path.join(base, "assets", "ferdlworks.ico")]
        for p in paths:
            if os.path.exists(p):
                try:
                    self.iconbitmap(p)
                    self.wm_iconbitmap(p)
                    self.logger.info(f"Icon gesetzt: {p}")
                    return
                except Exception:
                    continue

    # ===================== MENÜ =====================
    def _build_menu(self):
        mb = tk.Menu(self, font=("Segoe UI", 13))
        datei = tk.Menu(mb, tearoff=False, font=("Segoe UI", 10))
        datei.add_command(label="Neue Rechnung", command=self._new_doc_prompt, accelerator="Strg+N")
        datei.add_command(label="Speichern", command=self._save_doc, accelerator="Strg+S")
        datei.add_separator()
        datei.add_command(label="Rechnungen...", command=self._open_doc_overview, accelerator="Strg+D")
        datei.add_separator()
        datei.add_command(label="Einstellungen...", command=self._open_settings, accelerator="Strg+E")
        datei.add_separator()
        backup = tk.Menu(datei, tearoff=False, font=("Segoe UI", 10))
        backup.add_command(label="Datensicherung erstellen...", command=self._backup_data)
        backup.add_command(label="Datensicherung wiederherstellen...", command=self._restore_data)
        backup.add_separator()
        backup.add_command(label="Google Drive Backup...", command=self._cloud_gdrive)
        backup.add_command(label="Microsoft OneDrive Backup...", command=self._cloud_onedrive)
        datei.add_cascade(label="Backup", menu=backup)
        datei.add_separator()
        datei.add_command(label="Beenden", command=self._on_close, accelerator="Strg+Q")
        mb.add_cascade(label="Datei", menu=datei)

        verw = tk.Menu(mb, tearoff=False, font=("Segoe UI", 10))
        verw.add_command(label="Kundenverwaltung...", command=self._open_customer_mgmt)
        verw.add_separator()
        verw.add_command(label="Materialverwaltung...", command=self._open_material_mgmt)
        verw.add_command(label="Material importieren...", command=self._import_materials)
        verw.add_command(label="Material-Vorlage \u00f6ffnen...", command=self._open_material_template)
        verw.add_separator()
        verw.add_command(label="Werkzeugverwaltung...", command=self._open_tool_mgmt)
        verw.add_separator()
        verw.add_command(label="Texteverwaltung...", command=self._open_text_mgmt)
        mb.add_cascade(label="Verwaltung", menu=verw)

        hilfe = tk.Menu(mb, tearoff=False, font=("Segoe UI", 10))
        hilfe.add_command(label="Auf Updates prüfen...", command=self._check_update, accelerator="Strg+U")
        hilfe.add_separator()
        hilfe.add_command(label="Logdatei öffnen", command=self._open_log)
        hilfe.add_separator()
        hilfe.add_command(label="Info...", command=self._show_info)
        hilfe.add_separator()
        hilfe.add_command(label="Deinstallieren...", command=self._uninstall)
        mb.add_cascade(label="Hilfe", menu=hilfe)
        self.configure(menu=mb)
        self.bind_all("<Control-n>", lambda e: self._new_doc_prompt())
        self.bind_all("<Control-d>", lambda e: self._open_doc_overview())
        self.bind_all("<Control-s>", lambda e: self._save_doc())
        self.bind_all("<Control-q>", lambda e: self._on_close())
        self.bind_all("<Control-e>", lambda e: self._open_settings())
        self.bind_all("<Control-u>", lambda e: self._check_update())

    # ===================== UI =====================
    def _build_ui(self):
        main = ctk.CTkFrame(self)
        main.pack(fill="both", expand=True, padx=5, pady=5)

        # --- Kunde (Entry + Dropdown) ---
        cust = ctk.CTkFrame(main, corner_radius=6)
        cust.pack(fill="x", padx=8, pady=(4, 1))
        
        # Top row: Label + Entry + Buttons
        cust_top = ctk.CTkFrame(cust, fg_color="transparent")
        cust_top.pack(fill="x", padx=10, pady=(8, 0))
        ctk.CTkLabel(cust_top, text="Kunde:", font=("Segoe UI", 13, "bold"),
                     text_color=("#8b0000", "#8b0000"), width=70, anchor="w").pack(side="left")
        self.cust_entry = ctk.CTkEntry(cust_top, width=500)
        self.cust_entry._placeholder_text = "Kunde eingeben..."
        self.cust_entry.bind("<KeyRelease>", lambda e: self._filter_customers() if e.keysym not in ("Up", "Down", "Return", "Escape") else None)
        self.cust_entry.bind("<FocusIn>", lambda e: self._on_placeholder_in(self.cust_entry))
        self.cust_entry.bind("<FocusOut>", lambda e: self._on_placeholder_out(self.cust_entry))
        self.cust_entry.pack(side="left", padx=5)
        # Buttons Neu / Bearbeiten
        self.cust_btn_new = ctk.CTkButton(cust_top, text="Neu", width=60, 
                                          command=self._new_customer_from_main,
                                          fg_color="#5c0000", hover_color="#8b0000")
        self.cust_btn_new.pack(side="left", padx=5)
        self.cust_btn_edit = ctk.CTkButton(cust_top, text="Bearbeiten", width=80, 
                                           command=self._edit_customer_from_main, state="disabled",
                                           fg_color="#5c0000", hover_color="#8b0000")
        self.cust_btn_edit.pack(side="left", padx=5)
        self.doc_type_var = ctk.StringVar(value="RG")

        # Dropdown-Liste Kunde (schwebend über anderen Elementen)
        self._cust_dropdown_frame = tk.Frame(self, bg="#2a2a2a", highlightbackground="#555555", highlightthickness=1)
        self.cust_dropdown = tk.Listbox(self._cust_dropdown_frame, height=5,
                                        font=("Segoe UI", 13), exportselection=False,
                                        bg="#2a2a2a", fg="#e0e0e0", selectbackground="#8b0000",
                                        borderwidth=0, highlightthickness=0)
        self.cust_dropdown.pack(fill="x", padx=2, pady=2)
        self.cust_dropdown.bind("<<ListboxSelect>>", lambda e: self._pick_customer())
        # Tastatur-Navigation
        self.cust_entry.bind("<Down>", lambda e: self._nav_cust_down())
        self.cust_entry.bind("<Up>", lambda e: self._nav_cust_up())
        self.cust_entry.bind("<Return>", lambda e: self._enter_cust())
        self.cust_entry.bind("<Tab>", lambda e: self._enter_cust())
        self.cust_entry.bind("<Escape>", lambda e: self._do_hide_cust_dropdown())
        self._cust_data = []
        self._customer_id = None
        self._cust_dropdown_idx = -1

        # --- Artikel-Suche (Entry + Dropdown + Einheiten) ---
        art = ctk.CTkFrame(main, corner_radius=6)
        art.pack(fill="x", padx=8, pady=1)
        ctk.CTkLabel(art, text="Artikel:", font=("Segoe UI", 13, "bold"),
                     text_color=("#8b0000", "#8b0000"), width=70, anchor="w").pack(side="left", padx=(10, 0))
        self.art_entry = ctk.CTkEntry(art, width=500)
        self.art_entry._placeholder_text = "Werkzeug, Material oder Text eingeben..."
        self.art_entry.pack(side="left", padx=5, pady=4)
        self.art_entry.bind("<KeyRelease>", lambda e: self._search_articles() if e.keysym not in ("Up", "Down", "Return", "Escape") else None)
        self.art_entry.bind("<FocusIn>", lambda e: self._on_placeholder_in(self.art_entry))
        self.art_entry.bind("<FocusOut>", lambda e: self._on_placeholder_out(self.art_entry))

        # Kontext-Felder (abhängig vom ausgewählten Artikel-Typ, versteckt)
        self._art_mat_f = ctk.CTkFrame(art, fg_color="transparent")
        ctk.CTkLabel(self._art_mat_f, text="L:", font=("Segoe UI", 13)).pack(side="left")
        self.dl_length = ctk.CTkEntry(self._art_mat_f, width=55)
        self.dl_length.pack(side="left", padx=2)
        self.dl_length.bind("<KeyRelease>", lambda e: self._calc_detail_qm())
        ctk.CTkLabel(self._art_mat_f, text="cm", font=("Segoe UI", 13, "bold"),
                     text_color=("#666666", "#888888")).pack(side="left", padx=(0, 4))
        ctk.CTkLabel(self._art_mat_f, text="B:", font=("Segoe UI", 13)).pack(side="left")
        self.dl_width = ctk.CTkEntry(self._art_mat_f, width=55)
        self.dl_width.pack(side="left", padx=2)
        self.dl_width.bind("<KeyRelease>", lambda e: self._calc_detail_qm())
        self.dl_width.bind("<Tab>", lambda e: (self._insert_article(), self.art_entry.focus_set()) or "break")
        ctk.CTkLabel(self._art_mat_f, text="cm", font=("Segoe UI", 13, "bold"),
                     text_color=("#666666", "#888888")).pack(side="left", padx=(0, 4))
        self.dl_qm_label = ctk.CTkLabel(self._art_mat_f, text="m\xb2:0,00", font=("Segoe UI", 13, "bold"),
                                        text_color=("#8b0000", "#8b0000"))
        self.dl_qm_label.pack(side="left", padx=2)

        self._art_qty_f = ctk.CTkFrame(art, fg_color="transparent")
        ctk.CTkLabel(self._art_qty_f, text="Anzahl:", font=("Segoe UI", 13)).pack(side="left")
        self.dl_qty = ctk.CTkEntry(self._art_qty_f, width=80)
        self.dl_qty.pack(side="left", padx=2)
        self.dl_qty.bind("<Tab>", lambda e: (self._insert_article(), self.art_entry.focus_set()) or "break")

        self._art_tool_f = ctk.CTkFrame(art, fg_color="transparent")
        ctk.CTkLabel(self._art_tool_f, text="Zeit:", font=("Segoe UI", 13)).pack(side="left")
        self.dl_time = ctk.CTkEntry(self._art_tool_f, width=55)
        self.dl_time.insert(0, "1")
        self.dl_time.pack(side="left", padx=2)
        self.dl_time.bind("<Tab>", lambda e: (self._insert_article(), self.art_entry.focus_set()) or "break")
        self.dl_time_unit = ctk.StringVar(value="h")
        ctk.CTkOptionMenu(self._art_tool_f, variable=self.dl_time_unit, values=["h", "min"],
                          width=55).pack(side="left", padx=2)

        self.art_insert_btn = ctk.CTkButton(art, text="Einfügen", command=self._insert_article,
                                            width=80, fg_color="#5c0000", hover_color="#8b0000")
        self.art_insert_btn.pack(side="right", padx=5)

        self.art_text_btn = ctk.CTkButton(art, text="Als Text eintragen", command=self._insert_as_text,
                                          width=140, fg_color="#555555", hover_color="#777777")
        self.art_text_btn.pack(side="right", padx=2)
        self.art_text_btn.pack_forget()  # erstmal unsichtbar

        # Dropdown-Liste Artikel (schwebend, tabellarisch)
        self._art_dropdown_frame = tk.Frame(self, bg="#2a2a2a", highlightbackground="#555555", highlightthickness=1)
        art_style = ttk.Style()
        art_style.theme_use("clam")
        art_style.configure("Art.Treeview", background="#2a2a2a", foreground="#e0e0e0",
                            fieldbackground="#2a2a2a", rowheight=24, font=("Segoe UI", 13))
        art_style.map("Art.Treeview", background=[("selected", "#8b0000")], foreground=[("selected", "#ffffff")])
        art_style.configure("Art.Treeview.Heading", font=("Segoe UI", 10))
        art_style.layout("Art.Treeview", [("Art.Treeview.treearea", {"sticky": "nswe"})])
        self.art_dropdown = ttk.Treeview(self._art_dropdown_frame, columns=("typ", "name"),
                                          show="headings", height=5, style="Art.Treeview",
                                          selectmode="browse")
        self.art_dropdown.heading("typ", text="Typ", anchor="w")
        self.art_dropdown.heading("name", text="Name", anchor="w")
        self.art_dropdown.column("typ", width=150, anchor="w", minwidth=80)
        self.art_dropdown.column("name", width=350, anchor="w", minwidth=200)
        self.art_dropdown.pack(fill="x", padx=2, pady=2)
        self.art_dropdown.bind("<<TreeviewSelect>>", lambda e: self._select_article())
        self.art_entry.bind("<Down>", lambda e: self._nav_art_down())
        self.art_entry.bind("<Up>", lambda e: self._nav_art_up())
        self.art_entry.bind("<Return>", lambda e: self._enter_art())
        self.art_entry.bind("<Tab>", lambda e: self._enter_art())
        self.art_entry.bind("<Escape>", lambda e: self._do_hide_art_dropdown())
        self._art_results = []
        self._selected_article = None
        self._art_dropdown_idx = -1
        self._hide_units()

        # --- Positionen ---
        pos_frame = ctk.CTkFrame(main, corner_radius=6)
        pos_frame.pack(fill="both", expand=True, padx=8, pady=2)
        pos_header = ctk.CTkFrame(pos_frame, fg_color="transparent")
        pos_header.pack(fill="x", padx=10, pady=(6, 2))
        ctk.CTkLabel(pos_header, text="Positionen", font=("Segoe UI", 13, "bold"),
                     text_color=("#8b0000", "#8b0000")).pack(side="left")
        self.doc_status_label = ctk.CTkLabel(pos_header, text="", font=("Segoe UI", 13),
                                             text_color=("#666666", "#888888"))
        self.doc_status_label.pack(side="right")
        cols = ("pos", "beschreibung", "menge", "einheit", "ep", "gesamt")
        heads = {"pos": "Pos.", "beschreibung": "Beschreibung", "menge": "Menge",
                 "einheit": "Einheit", "ep": "EP", "gesamt": "Gesamt"}
        widths = {"pos": 40, "beschreibung": 350, "menge": 60, "einheit": 60, "ep": 80, "gesamt": 90}
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Treeview", background="#2a2a2a", foreground="#e0e0e0",
                        fieldbackground="#2a2a2a", rowheight=26, font=("Segoe UI", 13))
        style.map("Treeview", background=[("selected", "#8b0000")], foreground=[("selected", "#ffffff")])
        style.configure("Treeview.Heading", background="#5c0000", foreground="#ffffff",
                        font=("Segoe UI", 11, "bold"), relief="flat")
        style.map("Treeview.Heading", background=[("active", "#8b0000")])
        style.layout("Treeview", [("Treeview.treearea", {"sticky": "nswe"})])
        style.layout("Custom.Vertical.TScrollbar",
                     [("Vertical.Scrollbar.trough", {"sticky": "ns", "children":
                       [("Vertical.Scrollbar.thumb", {"sticky": "nswe"})]})])
        style.configure("Custom.Vertical.TScrollbar", background="#444444",
                        troughcolor="#2a2a2a", arrowcolor="#444444",
                        bordercolor="#2a2a2a", relief="flat", width=12)
        self.pos_tree = ttk.Treeview(pos_frame, columns=cols, show="headings", height=5)
        for c in cols:
            self.pos_tree.heading(c, text=heads[c], anchor="w" if c == "beschreibung" else "e")
            self.pos_tree.column(c, width=widths[c], minwidth=30, anchor="w" if c == "beschreibung" else "e")
        self.pos_tree.bind("<Delete>", lambda e: self._remove_selected_position())
        self.pos_tree.bind("<Double-1>", lambda e: self._edit_position())
        self.pos_tree.bind("<Button-3>", self._pos_context_menu)
        self.pos_tree.bind("<Button-2>", self._pos_context_menu)
        self._drag_data = {"item": None, "start_y": 0}
        self.pos_tree.bind("<Button-1>", self._drag_start, add="+")
        self.pos_tree.bind("<B1-Motion>", self._drag_motion)
        self.pos_tree.bind("<ButtonRelease-1>", self._drag_drop, add="+")
        self._pos_vsb = ttk.Scrollbar(pos_frame, orient="vertical", style="Custom.Vertical.TScrollbar",
                                       command=self.pos_tree.yview)
        self.pos_tree.configure(yscrollcommand=self._update_vsb)
        self.pos_tree.pack(fill="both", expand=True, padx=10, pady=2)

        # === 2x2 Grid unten (Notizen, Datum, Checkboxen, Buttons, Totalisierung) ===
        bottom = ctk.CTkFrame(main, corner_radius=6)
        bottom.pack(fill="x", padx=8, pady=(2, 0))
        bottom.grid_columnconfigure(0, weight=60)
        bottom.grid_columnconfigure(1, weight=40)
        bottom.grid_rowconfigure(0, weight=1)
        bottom.grid_rowconfigure(1, weight=1)

        # --- Top-Left: Notizen ---
        tl = ctk.CTkFrame(bottom, fg_color="transparent")
        tl.grid(row=0, column=0, sticky="nsew", padx=(0, 2), pady=(0, 2))
        nz = ctk.CTkFrame(tl, fg_color="transparent")
        nz.pack(anchor="w", padx=4, pady=(4, 0))
        ctk.CTkLabel(nz, text="Notiz:", font=("Segoe UI", 13, "bold"),
                     text_color="#8b0000").pack(side="left")
        self.print_note_var = ctk.BooleanVar(value=True)
        ctk.CTkCheckBox(nz, text="Auf Rechnung drucken unterhalb der Positionen", variable=self.print_note_var,
                        font=("Segoe UI", 13)).pack(side="left", padx=(6, 0))
        self.doc_note = ctk.CTkTextbox(tl, height=52, border_width=2)
        self.doc_note.pack(fill="x", padx=4, pady=(2, 4))
        ctk.CTkLabel(tl, text="Notiz (intern):", font=("Segoe UI", 13, "bold"),
                     text_color="#8b0000").pack(anchor="w", padx=4, pady=(0, 0))
        self.doc_internal_note = ctk.CTkTextbox(tl, height=52, border_width=2)
        self.doc_internal_note.pack(fill="x", padx=4, pady=(2, 4))

        # --- Top-Right: Datum + Zahlungsziel ---
        tr = ctk.CTkFrame(bottom, fg_color="transparent")
        tr.grid(row=0, column=1, sticky="nsew", padx=(2, 0), pady=(0, 2))
        ctk.CTkLabel(tr, text="Rechnungsdatum:", font=("Segoe UI", 13, "bold"),
                     text_color="#8b0000").pack(anchor="w", padx=6, pady=(6, 2))
        dr = ctk.CTkFrame(tr, fg_color="transparent")
        dr.pack(fill="x", padx=6)
        self.doc_date_var = ctk.StringVar(value=datetime.now().strftime("%d.%m.%Y"))
        self.doc_date_entry = ctk.CTkEntry(dr, width=85, textvariable=self.doc_date_var)
        self.doc_date_entry.pack(side="left")
        self.doc_date_var.trace_add("write", lambda *a: self._recalc_due_date())
        ctk.CTkButton(dr, text="\u2630", width=28, command=self._open_calendar,
                      fg_color="#555555").pack(side="left", padx=2)
        ctk.CTkButton(dr, text="Heute", width=45, command=lambda: self.doc_date_var.set(
            datetime.now().strftime("%d.%m.%Y")), fg_color="#555555").pack(side="left")
        zr = ctk.CTkFrame(tr, fg_color="transparent")
        zr.pack(fill="x", padx=6, pady=(6, 2))
        settings = self.db.settings_get_all()
        default_payment = int(settings.get("payment_term", "30"))
        ctk.CTkLabel(zr, text="Zahlungsziel:", font=("Segoe UI", 13, "bold"),
                     text_color="#8b0000").pack(side="left")
        self.payment_term_var = ctk.StringVar(value=str(default_payment))
        self.payment_term_entry = ctk.CTkEntry(zr, width=50, textvariable=self.payment_term_var)
        self.payment_term_entry.pack(side="left", padx=(2, 0))
        self.payment_term_var.trace_add("write", lambda *a: self._recalc_due_date())
        ctk.CTkButton(zr, text="\u25b2", width=25, command=lambda: self._adj_payment(1),
                      fg_color="#555555").pack(side="left", padx=1)
        ctk.CTkButton(zr, text="\u25bc", width=25, command=lambda: self._adj_payment(-1),
                      fg_color="#555555").pack(side="left")
        ctk.CTkLabel(zr, text="Zu bezahlen bis:", font=("Segoe UI", 13, "bold"),
                     text_color="#8b0000").pack(side="left", padx=(10, 0))
        self.due_date_var = ctk.StringVar()
        self.due_date_entry = ctk.CTkEntry(zr, width=105, textvariable=self.due_date_var,
                                           state="readonly")
        self.due_date_entry.pack(side="left", padx=(2, 0))
        self._recalc_due_date()

        # --- Bottom-Left: Checkboxen + Buttons ---
        bl = ctk.CTkFrame(bottom, fg_color="transparent")
        bl.grid(row=1, column=0, sticky="nsew", padx=(0, 2), pady=(2, 0))
        bl.grid_rowconfigure(0, weight=1)
        bl.grid_rowconfigure(1, weight=0)
        top_section = ctk.CTkFrame(bl, fg_color="transparent")
        top_section.grid(row=0, column=0, sticky="nsew")
        merge_frame = ctk.CTkFrame(top_section, fg_color="transparent")
        merge_frame.pack(fill="x", padx=6, pady=(4, 0), anchor="w")
        self.merge_tools_var = ctk.BooleanVar(value=True)
        self.merge_tools_cb = ctk.CTkCheckBox(merge_frame, text="Alle Werkzeugpositionen zusammenfassen zu:",
                                              variable=self.merge_tools_var,
                                              command=self._on_merge_tools_toggle,
                                              font=("Segoe UI", 13))
        self.merge_tools_cb.pack(side="left")
        self.merge_tool_name_var = ctk.StringVar(value="Arbeit")
        self.merge_tool_name_entry = ctk.CTkEntry(merge_frame, width=120, textvariable=self.merge_tool_name_var)
        self.merge_tool_name_entry.pack(side="left", padx=(4, 0))
        round_frame = ctk.CTkFrame(top_section, fg_color="transparent")
        round_frame.pack(fill="x", padx=6, pady=(1, 2))
        rfi = ctk.CTkFrame(round_frame, fg_color="transparent")
        rfi.pack(padx=(20, 0), anchor="w")
        self.round_tools_var = ctk.BooleanVar(value=True)
        self.round_tools_cb = ctk.CTkCheckBox(rfi, text="Auf ganze 10\u20ac aufrunden",
                                              variable=self.round_tools_var,
                                              command=self._on_round_tools_toggle,
                                              font=("Segoe UI", 13))
        self.round_tools_cb.pack(side="left")
        btn_frame = ctk.CTkFrame(bl, fg_color="transparent")
        btn_frame.grid(row=1, column=0, sticky="sew", padx=6, pady=(2, 6))
        for text, cmd in [("Speichern", self._save_doc),
                          ("PDF \u00f6ffnen", self._save_pdf),
                          ("PDF Ordner", self._save_pdf_folder),
                          ("E-Mail", self._send_email_doc),
                          ("Drucken", self._print_doc),
                          ("L\u00f6schen", self._delete_doc)]:
            fg = "#5c0000" if "L\u00f6schen" in text else "#8b0000"
            width = 95 if text == "PDF \u00f6ffnen" else 80
            ctk.CTkButton(btn_frame, text=text, command=cmd, width=width,
                          fg_color=fg, hover_color="#b22222",
                          font=("Segoe UI", 13)).pack(side="left", padx=2)

        # --- Bottom-Right: Totalisierung ---
        br = ctk.CTkFrame(bottom, fg_color="transparent")
        br.grid(row=1, column=1, sticky="nsew", padx=(2, 0), pady=(2, 0))
        self._build_totals(br)

        # --- Status Bar ---
        self._build_statusbar()

    def _on_placeholder_in(self, entry, ph=None):
        if getattr(entry, '_ph_active', False):
            entry.delete(0, "end")
            entry._ph_active = False

    def _on_placeholder_out(self, entry):
        if not entry.get().strip():
            ph = getattr(entry, '_placeholder_text', None)
            if ph:
                entry.delete(0, "end")
                entry.insert(0, ph)
                entry._ph_active = True

    # ===================== KUNDEN-DROPDOWN =====================
    def _on_global_click(self, event):
        w = self.winfo_containing(event.x_root, event.y_root)
        if self._cust_dropdown_frame.winfo_viewable():
            if w not in (self.cust_entry, self.cust_dropdown, self._cust_dropdown_frame) \
               and getattr(w, 'master', None) not in (self._cust_dropdown_frame,):
                self._do_hide_cust_dropdown()
        if self._art_dropdown_frame.winfo_viewable():
            if w not in (self.art_entry, self.art_dropdown, self._art_dropdown_frame) \
               and getattr(w, 'master', None) not in (self._art_dropdown_frame,):
                self._do_hide_art_dropdown()

    def _set_cust_display(self, cust):
        parts = [cust.get("company") or f"{cust.get('last_name', '')} {cust.get('first_name', '')}".strip()]
        if cust.get("street"):
            parts.append(cust["street"])
        zc = " ".join(filter(None, [cust.get("zip", ""), cust.get("city", "")]))
        if zc:
            parts.append(zc)
        self.cust_entry.delete(0, "end")
        self.cust_entry.insert(0, ", ".join(parts))
        self.cust_entry._ph_active = False

    def _filter_customers(self):
        query = self.cust_entry.get().strip()
        self._cust_data = self.db.customer_search(query) if query else []
        self.cust_dropdown.delete(0, "end")
        for c in self._cust_data:
            name = c.get("company") or f"{c.get('last_name', '')} {c.get('first_name', '')}".strip()
            orts = c.get("city", "")
            self.cust_dropdown.insert("end", f"{name}  ({orts})" if orts else name)
        if query and self._cust_data:
            self._show_cust_dropdown()
        else:
            self._do_hide_cust_dropdown()

    def _show_cust_dropdown(self):
        self._cust_dropdown_idx = -1
        x = self.cust_entry.winfo_rootx() - self.winfo_rootx()
        y = self.cust_entry.winfo_rooty() - self.winfo_rooty() + self.cust_entry.winfo_height()
        self._cust_dropdown_frame.place(x=x, y=y, width=500, anchor="nw")

    def _do_hide_cust_dropdown(self):
        self._cust_dropdown_frame.place_forget()
        self.cust_dropdown.selection_clear(0, "end")

    def _nav_cust_down(self, event=None):
        if not self._cust_data or not self._cust_dropdown_frame.winfo_viewable():
            return
        n = len(self._cust_data)
        if self._cust_dropdown_idx >= n - 1:
            return
        self._cust_dropdown_idx += 1
        idx = self._cust_dropdown_idx
        self._nav_arrow_cust = True
        self.cust_dropdown.selection_clear(0, "end")
        self.cust_dropdown.selection_set(idx)
        self.cust_dropdown.activate(idx)
        self.cust_dropdown.see(idx)
        self.after_idle(lambda: setattr(self, '_nav_arrow_cust', False))

    def _nav_cust_up(self, event=None):
        if not self._cust_data or not self._cust_dropdown_frame.winfo_viewable():
            return
        if self._cust_dropdown_idx <= 0:
            return
        self._cust_dropdown_idx -= 1
        idx = self._cust_dropdown_idx
        self._nav_arrow_cust = True
        self.cust_dropdown.selection_clear(0, "end")
        self.cust_dropdown.selection_set(idx)
        self.cust_dropdown.activate(idx)
        self.cust_dropdown.see(idx)
        self.after_idle(lambda: setattr(self, '_nav_arrow_cust', False))

    def _enter_cust(self, event=None):
        if not self._cust_data:
            return
        idx = self._cust_dropdown_idx
        if idx < 0 or idx >= len(self._cust_data):
            idx = 0
        self._cust_dropdown_idx = idx
        self._customer_id = self._cust_data[idx]["id"]
        self._set_cust_display(self._cust_data[idx])
        self.cust_btn_edit.configure(state="normal")
        self._do_hide_cust_dropdown()

    def _nav_art_down(self, event=None):
        if not self._art_results or not self._art_dropdown_frame.winfo_viewable():
            return
        n = len(self._art_results)
        if self._art_dropdown_idx >= n - 1:
            return
        self._art_dropdown_idx += 1
        idx = str(self._art_dropdown_idx)
        self._nav_arrow_art = True
        self.art_dropdown.selection_set()
        self.art_dropdown.selection_set(idx)
        self.art_dropdown.focus(idx)
        self.art_dropdown.see(idx)
        self.after_idle(lambda: setattr(self, '_nav_arrow_art', False))

    def _nav_art_up(self, event=None):
        if not self._art_results or not self._art_dropdown_frame.winfo_viewable():
            return
        if self._art_dropdown_idx <= 0:
            return
        self._art_dropdown_idx -= 1
        idx = str(self._art_dropdown_idx)
        self._nav_arrow_art = True
        self.art_dropdown.selection_set()
        self.art_dropdown.selection_set(idx)
        self.art_dropdown.focus(idx)
        self.art_dropdown.see(idx)
        self.after_idle(lambda: setattr(self, '_nav_arrow_art', False))

    def _enter_art(self, event=None):
        if not self._art_results:
            return
        idx = self._art_dropdown_idx
        if idx < 0 or idx >= len(self._art_results):
            idx = 0
        self._art_dropdown_idx = idx
        self._selected_article = self._art_results[idx]
        self.art_entry.delete(0, "end")
        self.art_entry.insert(0, self._selected_article["name"])
        self.art_entry._ph_active = False
        self._do_hide_art_dropdown()
        if self._selected_article["item_type"] == "Material":
            unit = self._selected_article.get("price_unit", "")
            if self._is_qm_unit(unit):
                self._show_mat_qm()
                self.dl_length.focus_set()
            else:
                self._show_mat_qty()
                self.dl_qty.focus_set()
        elif self._selected_article["item_type"] == "Werkzeug":
            self._show_tool_units()
            self.dl_time.focus_set()
        else:
            self._hide_units()

    def _pick_customer(self):
        if self._nav_arrow_cust:
            return
        sel = self.cust_dropdown.curselection()
        if not sel:
            return
        idx = sel[0]
        if idx < len(self._cust_data):
            self._cust_dropdown_idx = idx
            self._customer_id = self._cust_data[idx]["id"]
            self._set_cust_display(self._cust_data[idx])
            self._do_hide_cust_dropdown()

    def _get_selected_customer_id(self):
        return self._customer_id

    def _new_customer_from_main(self):
        dlg = CustomerDialog(self)
        self.wait_window(dlg)
        if dlg.result:
            self._customer_id = dlg.result["id"]
            self._set_cust_display(dlg.result)

    def _edit_customer_from_main(self):
        cid = self._get_selected_customer_id()
        if not cid:
            messagebox.showinfo("Hinweis", "Kein Kunde ausgewählt.")
            return
        dlg = CustomerDialog(self, customer_id=cid)
        self.wait_window(dlg)
        if dlg.result:
            self._set_cust_display(dlg.result)

    # ===================== ARTIKEL-DROPDOWN =====================
    def _search_articles(self):
        query = self.art_entry.get().strip()
        self._art_results = self.db.combined_search(query) if query else []
        for row in self.art_dropdown.get_children():
            self.art_dropdown.delete(row)
        for idx, item in enumerate(self._art_results):
            self.art_dropdown.insert("", "end", iid=str(idx), values=(item["item_type"], item["name"]))
        if query and self._art_results:
            self._show_art_dropdown()
            self.art_text_btn.pack_forget()
        elif query and not self._art_results:
            self._do_hide_art_dropdown()
            self.art_text_btn.pack(side="right", padx=2)
        else:
            self._do_hide_art_dropdown()
            self.art_text_btn.pack_forget()

    def _show_art_dropdown(self):
        self._art_dropdown_idx = -1
        x = self.art_entry.winfo_rootx() - self.winfo_rootx()
        y = self.art_entry.winfo_rooty() - self.winfo_rooty() + self.art_entry.winfo_height()
        self._art_dropdown_frame.place(x=x, y=y, width=500, anchor="nw")

    def _do_hide_art_dropdown(self):
        self._art_dropdown_frame.place_forget()
        self.art_dropdown.selection_set()

    def _select_article(self):
        if self._nav_arrow_art:
            return
        sel = self.art_dropdown.selection()
        if not sel:
            return
        idx = int(sel[0])
        if idx >= len(self._art_results):
            return
        self._art_dropdown_idx = idx
        self._selected_article = self._art_results[idx]
        self.art_entry.delete(0, "end")
        self.art_entry.insert(0, self._selected_article["name"])
        self.art_entry._ph_active = False
        self._do_hide_art_dropdown()
        if self._selected_article["item_type"] == "Material":
            unit = self._selected_article.get("price_unit", "")
            if self._is_qm_unit(unit):
                self._show_mat_qm()
                self.dl_length.focus_set()
            else:
                self._show_mat_qty()
                self.dl_qty.focus_set()
        elif self._selected_article["item_type"] == "Werkzeug":
            self._show_tool_units()
            self.dl_time.focus_set()
        else:
            self._hide_units()

    @staticmethod
    def _is_qm_unit(unit):
        return unit and unit.strip().lower() in ("m2", "m\u00b2", "qm")

    def _show_mat_qm(self):
        self._art_tool_f.pack_forget()
        self._art_qty_f.pack_forget()
        self._art_mat_f.pack(side="left", padx=(5, 0))
        self.art_insert_btn.configure(text="Übernehmen" if self._editing_pos_idx is not None else "Einfügen")
        self.dl_length.delete(0, "end")
        self.dl_width.delete(0, "end")

    def _show_mat_qty(self):
        self._art_mat_f.pack_forget()
        self._art_tool_f.pack_forget()
        self._art_qty_f.pack(side="left", padx=(5, 0))
        self.art_insert_btn.configure(text="Übernehmen" if self._editing_pos_idx is not None else "Einfügen")
        self.dl_qty.delete(0, "end")
        self.dl_qty.insert(0, "1")

    def _show_tool_units(self):
        self._art_mat_f.pack_forget()
        self._art_tool_f.pack(side="left", padx=(5, 0))
        self.art_insert_btn.configure(text="Übernehmen" if self._editing_pos_idx is not None else "Einfügen")
        self.dl_time.delete(0, "end")
        self.dl_time.insert(0, "1")

    def _hide_units(self):
        self._art_mat_f.pack_forget()
        self._art_tool_f.pack_forget()
        self._art_qty_f.pack_forget()

    def _calc_detail_qm(self):
        try:
            length = float(self.dl_length.get().replace(",", ".")) / 100
            width = float(self.dl_width.get().replace(",", ".")) / 100
            qm = length * width
            self.dl_qm_label.configure(text=f"m\xb2:{qm:.2f}".replace(".", ","))
        except ValueError:
            self.dl_qm_label.configure(text="m\xb2:0,00")

    # ===================== EINFÜGEN / ÜBERNEHMEN =====================
    def _calc_tool_position(self, item):
        try:
            time_val = float(self.dl_time.get().replace(",", "."))
        except ValueError:
            time_val = 1
        display_unit = self.dl_time_unit.get()
        stored_unit = item.get("price_unit", "h")
        price = item.get("price", 0)
        rate_per_min = price / 60 if stored_unit == "h" else price
        minutes = time_val * 60 if display_unit == "h" else time_val
        total = minutes * rate_per_min
        price_per_display = rate_per_min * 60 if display_unit == "h" else rate_per_min
        unit_label = "Std." if display_unit == "h" else "Min."
        return time_val, unit_label, price_per_display, total, int(minutes)

    def _insert_as_text(self):
        text = self.art_entry.get().strip()
        if not text:
            return
        self._positions.append(PositionItem("text", None, text, 0, "", 0, 0))
        self._refresh_positions()
        self._hide_units()
        self.art_entry.delete(0, "end")
        self.art_text_btn.pack_forget()

    def _insert_article(self):
        if self._editing_pos_idx is not None:
            self._update_position()
        else:
            self._add_position()
        self._refresh_positions()
        self._hide_units()
        self._selected_article = None
        self.art_entry.delete(0, "end")
        self._on_placeholder_out(self.art_entry)
        self.after_idle(lambda: self.art_entry.focus_set())

    def _add_position(self):
        item = self._selected_article
        if not item:
            return
        if item["item_type"] == "Material":
            mat_unit = item.get("price_unit", "m\u00b2")
            if self._is_qm_unit(mat_unit):
                try:
                    length = float(self.dl_length.get().replace(",", ".")) if self.dl_length.get() else 0
                    width = float(self.dl_width.get().replace(",", ".")) if self.dl_width.get() else 0
                except ValueError:
                    length = width = 0
                price_m2 = item.get("price_per_m2", 0) or item.get("price", 0)
                if length > 0 and width > 0:
                    qm = (length / 100) * (width / 100)
                    desc = f"{item['name']} ({length:.0f}x{width:.0f}cm)"
                    total = qm * price_m2
                    self._positions.append(PositionItem(
                        "material", item["id"], desc, qm, mat_unit, price_m2, total,
                        {"length": length, "width": width, "qty": 1}
                    ))
                else:
                    self._positions.append(PositionItem(
                        "material", item["id"], item["name"], 1, mat_unit, price_m2, price_m2
                    ))
            else:
                try:
                    qty = float(self.dl_qty.get().replace(",", ".")) if self.dl_qty.get() else 0
                except ValueError:
                    qty = 1
                price = item.get("price", 0)
                total = qty * price
                desc = item['name']
                self._positions.append(PositionItem(
                    "material", item["id"], desc, qty, mat_unit, price, total,
                    {"qty": qty}
                ))
        elif item["item_type"] == "Werkzeug":
            time_val, unit_label, price_per, total, _ = self._calc_tool_position(item)
            desc = item["name"]
            self._positions.append(PositionItem(
                "tool", item["id"], desc, time_val, unit_label, price_per, total,
                {"price_unit": item.get("price_unit", "h"), "price": item.get("price", 0)}
            ))
        elif item["item_type"] == "Text":
            desc = item.get("content") or item["name"]
            self._positions.append(PositionItem(
                "text", item["id"], desc, 0, "", 0, 0
            ))

    def _update_position(self):
        item = self._selected_article
        if not item:
            return
        idx = self._editing_pos_idx
        if item["item_type"] == "Material":
            mat_unit = item.get("price_unit", "m\u00b2")
            if self._is_qm_unit(mat_unit):
                try:
                    length = float(self.dl_length.get().replace(",", ".")) if self.dl_length.get() else 0
                    width = float(self.dl_width.get().replace(",", ".")) if self.dl_width.get() else 0
                except ValueError:
                    length = width = 0
                price_m2 = item.get("price_per_m2", 0) or item.get("price", 0)
                if length > 0 and width > 0:
                    qm = (length / 100) * (width / 100)
                    desc = f"{item['name']} ({length:.0f}x{width:.0f}cm)"
                    total = qm * price_m2
                    self._positions[idx] = PositionItem(
                        "material", item["id"], desc, qm, mat_unit, price_m2, total,
                        {"length": length, "width": width, "qty": 1}
                    )
                else:
                    self._positions[idx] = PositionItem(
                        "material", item["id"], item["name"], 1, mat_unit, price_m2, price_m2
                    )
            else:
                try:
                    qty = float(self.dl_qty.get().replace(",", ".")) if self.dl_qty.get() else 0
                except ValueError:
                    qty = 1
                price = item.get("price", 0)
                total = qty * price
                desc = item['name']
                self._positions[idx] = PositionItem(
                    "material", item["id"], desc, qty, mat_unit, price, total,
                    {"qty": qty}
                )
        elif item["item_type"] == "Werkzeug":
            time_val, unit_label, price_per, total, _ = self._calc_tool_position(item)
            desc = item["name"]
            self._positions[idx] = PositionItem(
                "tool", item["id"], desc, time_val, unit_label, price_per, total,
                {"price_unit": item.get("price_unit", "h"), "price": item.get("price", 0)}
            )
        elif item["item_type"] == "Text":
            desc = item.get("content") or item["name"]
            self._positions[idx] = PositionItem(
                "text", item["id"], desc, 0, "", 0, 0
            )
        self._editing_pos_idx = None
        self.art_insert_btn.configure(text="Einfügen")

    # ===================== POSITIONEN =====================
    def _update_vsb(self, first, last):
        try:
            if float(first) <= 0.0 and float(last) >= 1.0:
                self._pos_vsb.pack_forget()
                return
        except ValueError:
            pass
        self._pos_vsb.pack(side="right", fill="y")
        self._pos_vsb.set(first, last)

    def _refresh_positions(self):
        for row in self.pos_tree.get_children():
            self.pos_tree.delete(row)
        for i, p in enumerate(self._positions):
            qty_str = f"{p.quantity:.2f}" if p.quantity != int(p.quantity) else str(int(p.quantity))
            if p.pos_type == "tool" and p.extra_data and "price_unit" in p.extra_data:
                ep_str = f"{p.extra_data['price']:.2f}\u20ac/{p.extra_data['price_unit']}"
            elif p.pos_type == "text":
                ep_str = ""
            else:
                ep_str = f"{p.price_per_unit:.2f}\u20ac/{p.unit.lower().replace('std.', 'h').replace('min.', 'min')}"
            if p.pos_type == "text":
                vals = ("", p.description, "", "", "", "")
            else:
                vals = (str(i + 1), p.description, qty_str, p.unit, ep_str, f"{p.total:.2f}\u20ac")
            tag = "even" if i % 2 == 0 else "odd"
            self.pos_tree.insert("", "end", iid=str(i), values=vals, tags=(tag,))
        self._recalc_totals()

    def _drag_start(self, event):
        iid = self.pos_tree.identify_row(event.y)
        if iid:
            self._drag_data["item"] = int(iid)
            self._drag_data["start_y"] = event.y

    def _drag_motion(self, event):
        iid = self.pos_tree.identify_row(event.y)
        if iid is not None and self._drag_data["item"] is not None:
            self.pos_tree.selection_set(iid)

    def _drag_drop(self, event):
        if self._drag_data["item"] is None:
            return
        target_iid = self.pos_tree.identify_row(event.y)
        if target_iid and int(target_iid) != self._drag_data["item"]:
            src = self._drag_data["item"]
            dst = int(target_iid)
            item = self._positions.pop(src)
            self._positions.insert(dst, item)
            self._refresh_positions()
            self.pos_tree.selection_set(str(dst))
        self._drag_data["item"] = None

    def _pos_context_menu(self, event):
        # select the clicked row
        iid = self.pos_tree.identify_row(event.y)
        if not iid:
            return
        self.pos_tree.selection_set(iid)
        menu = tk.Menu(self, tearoff=False, font=("Segoe UI", 10))
        menu.add_command(label="Bearbeiten", command=self._edit_position)
        menu.add_command(label="Löschen", command=self._remove_selected_position)
        menu.tk_popup(event.x_root, event.y_root)
        menu.grab_release()

    def _remove_selected_position(self):
        sel = self.pos_tree.selection()
        if sel:
            idx = int(sel[0])
            if 0 <= idx < len(self._positions):
                self._positions.pop(idx)
                self._refresh_positions()

    def _edit_position(self):
        sel = self.pos_tree.selection()
        if not sel:
            return
        idx = int(sel[0])
        if idx >= len(self._positions):
            return
        pos = self._positions[idx]
        self._editing_pos_idx = idx
        self.art_entry.delete(0, "end")
        self.art_entry.insert(0, pos.description)
        # Artikel per Datenbank ermitteln, Fallback auf minimale Daten
        found = None
        if pos.ref_id:
            if pos.pos_type == "material":
                found = self.db.material_get(pos.ref_id)
                if found:
                    found["item_type"] = "Material"
                    found["price_per_m2"] = found.get("price_per_m2", 0) or found.get("price", 0)
            else:
                found = self.db.tool_get(pos.ref_id)
                if found:
                    found["item_type"] = "Werkzeug"
                    found["price"] = found.get("price", 0)
                    found["price_per_m2"] = 0
        if not found:
            _t = "Text" if pos.pos_type == "text" else ("Material" if pos.pos_type == "material" else "Werkzeug")
            found = {"id": pos.ref_id or 0, "name": pos.description,
                     "item_type": _t, "price": 0, "price_per_m2": 0, "price_unit": "h"}
        self._selected_article = found
        if pos.pos_type == "material":
            unit = found.get("price_unit", pos.unit or "")
            if self._is_qm_unit(unit):
                self._show_mat_qm()
                ed = pos.extra_data or {}
                if ed.get("length"):
                    self.dl_length.delete(0, "end")
                    self.dl_length.insert(0, str(ed["length"]))
                if ed.get("width"):
                    self.dl_width.delete(0, "end")
                    self.dl_width.insert(0, str(ed["width"]))
                self._calc_detail_qm()
            else:
                self._show_mat_qty()
                self.dl_qty.delete(0, "end")
                self.dl_qty.insert(0, str(int(pos.quantity)))
        elif pos.pos_type == "tool":
            self._show_tool_units()
            self.dl_time.delete(0, "end")
            self.dl_time.insert(0, str(int(pos.quantity)))
            self.dl_time_unit.set("h" if pos.unit == "Std." else "min")
        else:
            self._hide_units()
        self.art_insert_btn.configure(text="Übernehmen")

    # ===================== SUMMEN =====================
    def _get_effective_net(self):
        tool_sum = sum(p.total for p in self._positions if p.pos_type == "tool")
        other_sum = sum(p.total for p in self._positions if p.pos_type != "tool")
        if self.round_tools_var.get():
            tool_sum = math.ceil(tool_sum / 10) * 10
        return tool_sum + other_sum

    def _recalc_totals(self):
        settings = self.db.settings_get_all()
        tax_rate = float(settings.get("tax_rate", "19"))
        total_net = self._get_effective_net()
        try:
            discount_val = float(self.discount_var.get().replace(",", "."))
        except ValueError:
            discount_val = 0
        is_percent = self.discount_type_var.get() == "%"
        rabatt = total_net * discount_val / 100 if is_percent else discount_val
        if not is_percent and discount_val > 0:
            self.discount_var.set(f"{discount_val:.2f}".replace(".", ","))
        netto_nach_rabatt = max(0, total_net - rabatt)
        total_tax = netto_nach_rabatt * tax_rate / 100
        total_gross = netto_nach_rabatt + total_tax
        self._sum_labels["netto"].configure(text=f"{total_net:.2f}\u20ac".replace(".", ","))
        self._sum_labels["rabatt"].configure(
            text=f"-{rabatt:.2f}\u20ac".replace(".", ",") if rabatt > 0 else "0,00\u20ac")
        self._sum_labels["mwst"].configure(text=f"{total_tax:.2f}\u20ac".replace(".", ","))
        self._sum_labels["brutto"].configure(text=f"{total_gross:.2f}\u20ac".replace(".", ","))

    # ===================== TOTALISIERUNG =====================
    def _build_totals(self, parent):
        self._sum_labels = {}
        r = ctk.CTkFrame(parent, fg_color="transparent")
        r.pack(fill="x", padx=6, pady=(8, 2))
        ctk.CTkLabel(r, text="Rabatt:", font=("Segoe UI", 15, "bold"),
                     text_color="#8b0000", width=55, anchor="w").pack(side="left")
        self.discount_var = ctk.StringVar(value="0")
        self.discount_entry = ctk.CTkEntry(r, width=70, textvariable=self.discount_var)
        self.discount_entry.pack(side="left", padx=2)
        self.discount_entry.bind("<KeyRelease>", lambda e: self._recalc_totals())
        self.discount_type_var = ctk.StringVar(value="%")
        ctk.CTkOptionMenu(r, variable=self.discount_type_var, values=["%", "\u20ac"],
                          width=60, command=lambda v: self._recalc_totals()).pack(side="left")
        for text, key in [("Netto:", "netto"), ("Rabatt:", "rabatt"), ("MwSt:", "mwst"), ("Brutto:", "brutto")]:
            f = ctk.CTkFrame(parent, fg_color="transparent")
            f.pack(fill="x", padx=6, pady=1)
            ctk.CTkLabel(f, text=text, font=("Segoe UI", 17)).pack(side="left")
            lbl = ctk.CTkLabel(f, text="0,00 \u20ac", font=("Segoe UI", 17, "bold"),
                               text_color=("#8b0000", "#8b0000"), width=80, anchor="e")
            lbl.pack(side="right", padx=4)
            self._sum_labels[key] = lbl

    # ===================== DATUM / ZAHLUNGSZIEL =====================
    def _recalc_due_date(self):
        try:
            doc_date = datetime.strptime(self.doc_date_var.get(), "%d.%m.%Y")
        except (ValueError, TypeError):
            self.due_date_var.set("")
            return
        try:
            days = int(self.payment_term_var.get())
        except ValueError:
            days = 30
        due = doc_date + timedelta(days=days)
        self.due_date_var.set(due.strftime("%d.%m.%Y"))

    def _adj_payment(self, delta):
        try:
            cur = int(self.payment_term_var.get())
        except ValueError:
            cur = 30
        cur = max(0, cur + delta)
        self.payment_term_var.set(str(cur))

    def _on_merge_tools_toggle(self):
        if not self.merge_tools_var.get():
            self.round_tools_var.set(False)
        self._recalc_totals()

    def _on_round_tools_toggle(self):
        if self.round_tools_var.get():
            self.merge_tools_var.set(True)
        self._recalc_totals()

    # ===================== PDF ORDNER =====================
    def _save_pdf_folder(self):
        if getattr(sys, 'frozen', False):
            folder = os.path.join(os.environ.get("APPDATA", os.path.expanduser("~")), "FerdlWorks", "data", "pdfs")
        else:
            folder = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "pdfs")
        if os.path.isdir(folder):
            os.startfile(folder)
        else:
            messagebox.showinfo("PDF Ordner", f"PDF-Ordner existiert noch nicht:\n{folder}")

    # ===================== STATUSBAR =====================
    def _build_statusbar(self):
        bar = ctk.CTkFrame(self, height=24, corner_radius=0, fg_color=("#e0e0e0", "#1a1a1a"))
        bar.pack(fill="x", side="bottom")
        ctk.CTkLabel(bar, text="SondereggerSoftware", font=("Segoe UI", 10),
                     text_color=("#555555", "#888888")).pack(side="left", padx=10)
        ctk.CTkLabel(bar, text=f"v{VERSION}", font=("Segoe UI", 10),
                     text_color=("#555555", "#888888")).pack(side="right", padx=10)

    # ===================== DOKUMENT-LOGIK =====================
    def _new_doc_prompt(self):
        if self._positions:
            doc_type = {"RG": "Rechnung", "LS": "Lieferschein"}.get(self.doc_type_var.get(), "Rechnung")
            ans = messagebox.askyesnocancel("Nicht gespeicherte Änderungen",
                                            f"Möchten Sie die aktuelle {doc_type} speichern?")
            if ans is None:
                return
            if ans:
                self._save_doc()
        self._new_doc()

    def _new_doc(self):
        self._current_doc_id = None
        self._positions.clear()
        self._refresh_positions()
        self.cust_entry.delete(0, "end")
        self.cust_entry._ph_active = False
        self._on_placeholder_out(self.cust_entry)
        self.art_entry.delete(0, "end")
        self.art_entry._ph_active = False
        self._on_placeholder_out(self.art_entry)
        self.doc_note.delete("1.0", "end")
        self.doc_internal_note.delete("1.0", "end")
        self.doc_date_var.set(datetime.now().strftime("%d.%m.%Y"))
        self.discount_var.set("0")
        self.doc_type_var.set("RG")
        self._customer_id = None
        self.cust_btn_edit.configure(state="disabled")
        self._cust_data = []
        self._editing_pos_idx = None
        self._hide_units()
        self._do_hide_cust_dropdown()
        self._do_hide_art_dropdown()
        doc_type = {"RG": "Rechnung", "LS": "Lieferschein"}.get(self.doc_type_var.get(), "Rechnung")
        self.doc_status_label.configure(text=f"Neue {doc_type}, nicht gespeichert")

    def _get_doc_data(self):
        settings = self.db.settings_get_all()
        tax_rate = float(settings.get("tax_rate", "19"))
        orig_net = self._get_effective_net()
        try:
            discount_val = float(self.discount_var.get().replace(",", "."))
        except ValueError:
            discount_val = 0
        is_percent = self.discount_type_var.get() == "%"
        if discount_val > 0:
            rabatt = orig_net * discount_val / 100 if is_percent else discount_val
            net_after = max(0, orig_net - rabatt)
        else:
            rabatt = 0
            net_after = orig_net
        total_tax = net_after * tax_rate / 100
        total_gross = net_after + total_tax
        return {
            "doc_type": self.doc_type_var.get(),
            "customer_id": self._get_selected_customer_id(),
            "date": datetime.now().strftime("%Y-%m-%d") if not self.doc_date_var.get().strip() else datetime.strptime(self.doc_date_var.get(), "%d.%m.%Y").strftime("%Y-%m-%d"),
            "due_date": datetime.strptime(self.due_date_var.get(), "%d.%m.%Y").strftime("%Y-%m-%d") if self.due_date_var.get().strip() else "",
            "discount_type": "percent" if is_percent else "fixed",
            "discount_value": discount_val,
            "total_net": orig_net,
            "total_tax": total_tax,
            "total_gross": total_gross,
            "note": self.doc_note.get("1.0", "end-1c"),
            "internal_note": self.doc_internal_note.get("1.0", "end-1c"),
            "print_note": "1" if self.print_note_var.get() else "0",
            "merge_tools": self.merge_tools_var.get(),
            "merge_tool_name": self.merge_tool_name_var.get(),
            "round_tools": self.round_tools_var.get(),
        }

    def _do_save(self, silent=False):
        cid = self._get_selected_customer_id()
        if not cid:
            if not silent:
                messagebox.showwarning("Fehler", "Bitte wählen Sie einen Kunden aus.")
            return None
        if not self._positions:
            if not silent:
                messagebox.showwarning("Fehler", "Keine Positionen vorhanden.")
            return None
        data = self._get_doc_data()
        pos_data = []
        for p in self._positions:
            ed = p.extra_data or {}
            pos_data.append({
                "pos_type": p.pos_type,
                "ref_id": p.ref_id,
                "description": p.description,
                "quantity": p.quantity,
                "unit": p.unit,
                "price_per_unit": p.price_per_unit,
                "total": p.total,
                "orig_price": ed.get("price") or ed.get("orig_price", 0),
                "orig_price_unit": ed.get("price_unit") or ed.get("orig_price_unit", ""),
            })
        data["id"] = self._current_doc_id
        result = self.db.doc_save(data, pos_data)
        if result:
            self._current_doc_id = result["id"]
            self.logger.info(f"Dokument {result['doc_number']} gespeichert")
            if not silent:
                messagebox.showinfo("Gespeichert", f"{'Rechnung' if result['doc_type'] == 'RG' else 'Lieferschein'} {result['doc_number']} gespeichert.")
            self._load_doc(result["id"])
        return result

    def _save_doc(self):
        self._do_save(silent=False)

    def _load_doc(self, doc_id):
        doc = self.db.doc_get(doc_id)
        if not doc:
            return
        self._current_doc_id = doc["id"]
        self.doc_type_var.set(doc["doc_type"])
        self.doc_note.delete("1.0", "end")
        self.doc_note.insert("1.0", doc.get("note", ""))
        self.doc_internal_note.delete("1.0", "end")
        self.doc_internal_note.insert("1.0", doc.get("internal_note", ""))
        self.print_note_var.set(doc.get("print_note", "1") == "1")
        self.merge_tools_var.set(doc.get("merge_tools", "1") == "1")
        self.merge_tool_name_var.set(doc.get("merge_tool_name", "Werkzeug"))
        self.round_tools_var.set(doc.get("round_tools", "1") == "1")
        self.discount_var.set(str(doc.get("discount_value", "0")).replace(".", ","))
        self.discount_type_var.set("%" if doc.get("discount_type", "percent") == "percent" else "\u20ac")
        try:
            d = datetime.strptime(doc.get("date", ""), "%Y-%m-%d")
            self.doc_date_var.set(d.strftime("%d.%m.%Y"))
        except ValueError:
            self.doc_date_var.set(datetime.now().strftime("%d.%m.%Y"))
        try:
            dd = datetime.strptime(doc.get("due_date", ""), "%Y-%m-%d")
            self.due_date_var.set(dd.strftime("%d.%m.%Y"))
            days_diff = (dd - d).days if d else 30
            self.payment_term_var.set(str(max(0, days_diff)))
        except (ValueError, TypeError):
            self._recalc_due_date()
        customer = doc.get("customer")
        if customer:
            self._customer_id = customer["id"]
            self._set_cust_display(customer)
        self._positions.clear()
        for p in doc.get("positions", []):
            ed = {}
            if p.get("orig_price_unit"):
                ed = {"price_unit": p["orig_price_unit"], "price": float(p.get("orig_price", 0) or 0)}
            self._positions.append(PositionItem(
                p["pos_type"], p["ref_id"], p["description"],
                p["quantity"], p["unit"], p["price_per_unit"], p["total"], ed
            ))
        self._refresh_positions()
        self.doc_status_label.configure(text=doc["doc_number"])

    def _check_overdue(self):
        overdue = self.db.doc_get_overdue()
        if overdue:
            msg = f"{len(overdue)} Rechnung(en) sind überfällig!\n\nDiese anzeigen?"
            if messagebox.askyesno("Überfällige Rechnungen", msg):
                DocSearchDialog(self, overdue_only=True)

    def _open_doc_search(self):
        DocSearchDialog(self)

    def _delete_doc(self):
        if self._check_empty():
            return
        if messagebox.askyesno("Löschen", "Aktuelles Dokument endgültig löschen?\n\nDiese Aktion kann nicht rückgängig gemacht werden."):
            self.db.doc_delete(self._current_doc_id)
            self._new_doc()

    def _open_calendar(self):
        from tkinter import ttk
        top = ctk.CTkToplevel(self)
        top.title("Datum auswählen")
        top.geometry("280x250")
        top.transient(self)
        top.grab_set()
        top.resizable(False, False)
        now = datetime.now()
        m, y = now.month, now.year
        sel = self.doc_date_var.get()
        try:
            sd = datetime.strptime(sel, "%d.%m.%Y")
            cm, cy = sd.month, sd.year
        except:
            cm, cy = m, y

        def build(month, year):
            for w in top.winfo_children():
                w.destroy()
            hf = ctk.CTkFrame(top, fg_color="transparent")
            hf.pack(fill="x", padx=5, pady=5)
            ctk.CTkButton(hf, text="<", width=30, command=lambda: build(month-1 if month>1 else 12, year if month>1 else year-1)).pack(side="left")
            ctk.CTkLabel(hf, text=f"{month:02d}/{year}", font=("Segoe UI", 13, "bold")).pack(side="left", expand=True)
            ctk.CTkButton(hf, text=">", width=30, command=lambda: build(month+1 if month<12 else 1, year if month<12 else year+1)).pack(side="right")
            cf = ctk.CTkFrame(top, fg_color="transparent")
            cf.pack(padx=5, pady=2)
            days = ["Mo","Di","Mi","Do","Fr","Sa","So"]
            for i, d in enumerate(days):
                ctk.CTkLabel(cf, text=d, width=30, font=("Segoe UI", 9)).grid(row=0, column=i)
            import calendar
            cal = calendar.monthcalendar(year, month)
            for r, week in enumerate(cal, 1):
                for c, day in enumerate(week):
                    if day == 0:
                        ctk.CTkLabel(cf, text="", width=30).grid(row=r, column=c)
                    else:
                        btn = ctk.CTkButton(cf, text=str(day), width=30, height=25, fg_color="#555555",
                                            command=lambda d=day, mo=month, yr=year: pick(d, mo, yr))
                        btn.grid(row=r, column=c, padx=1, pady=1)

        def pick(day, month, year):
            self.doc_date_var.set(f"{day:02d}.{month:02d}.{year}")
            top.destroy()

        build(cm, cy)

    # ===================== PDF / E-MAIL / DRUCKEN =====================
    def _check_empty(self):
        if not self._positions:
            messagebox.showwarning("Hinweis", "Keine Einträge vorhanden.")
            return True
        return False

    def _save_pdf(self):
        if self._check_empty():
            return
        if not self._get_selected_customer_id():
            messagebox.showwarning("Fehler", "Bitte wählen Sie einen Kunden aus.")
            return
        if not self._do_save(silent=True):
            return
        doc = self.db.doc_get(self._current_doc_id)
        if not doc:
            messagebox.showerror("Fehler", "Dokument konnte nicht geladen werden.")
            return
        doc["app_name"] = APP_NAME
        try:
            path = generate_pdf(doc)
            if path:
                self.logger.info(f"PDF geöffnet: {path}")
                os.startfile(path)
            else:
                messagebox.showerror("Fehler", "PDF konnte nicht erstellt werden.")
        except Exception as ex:
            messagebox.showerror("Fehler", f"PDF-Fehler: {ex}")

    def _send_email_doc(self):
        if self._check_empty():
            return
        settings = self.db.settings_get_all()
        if not settings.get("smtp_host") or not settings.get("smtp_user"):
            messagebox.showwarning("E-Mail", "Bitte zuerst E-Mail-Einstellungen konfigurieren (Einstellungen > E-Mail).")
            return
        if not self._current_doc_id and not self._do_save(silent=True):
            return
        doc = self.db.doc_get(self._current_doc_id)
        if not doc:
            return
        customer = doc.get("customer", {})
        recipient = customer.get("email", "")
        if not recipient:
            messagebox.showwarning("E-Mail", "Kunde hat keine E-Mail-Adresse hinterlegt.")
            return
        if not messagebox.askyesno("E-Mail", "Dokument per E-Mail verschicken?"):
            return
        doc["app_name"] = APP_NAME
        pdf_path = generate_pdf(doc)
        # Variablen ersetzen
        def _fmt(v):
            return v.replace(".", ",") if isinstance(v, str) else f"{v:.2f}\u20ac".replace(".", ",")
        raw_subject = settings.get("email_subject", "Ihre Rechnung {rgnr}")
        raw_body = settings.get("email_body", "")
        vars_map = {
            "{vorname}": customer.get("first_name", ""),
            "{nachname}": customer.get("last_name", ""),
            "{rgnr}": doc.get("doc_number", ""),
            "{rgdat}": doc.get("date", ""),
            "{bezbisdatum}": doc.get("due_date", ""),
            "{bezbistage}": str(max(0, (datetime.strptime(doc.get("due_date", ""), "%Y-%m-%d") - datetime.strptime(doc.get("date", ""), "%Y-%m-%d")).days)) if doc.get("due_date") and doc.get("date") else "30",
            "{betrag}": f"{doc.get('total_gross', 0):.2f}\u20ac".replace(".", ","),
        }
        for key, val in vars_map.items():
            raw_subject = raw_subject.replace(key, str(val))
            raw_body = raw_body.replace(key, str(val))
        subject = raw_subject
        body = raw_body
        success, msg = send_email(recipient, subject, body, pdf_path)
        if success:
            messagebox.showinfo("E-Mail", msg)
        else:
            messagebox.showerror("Fehler", msg)

    def _print_doc(self):
        if self._check_empty():
            return
        printer = self.db.settings_get_all().get("printer_name", "")
        if not printer:
            messagebox.showwarning("Drucken", "Kein Drucker ausgewählt. Bitte wählen Sie einen Drucker in den Einstellungen.")
            return
        if not self._current_doc_id and not self._do_save(silent=True):
            return
        doc = self.db.doc_get(self._current_doc_id)
        if not doc:
            return
        doc["app_name"] = APP_NAME
        pdf_path = generate_pdf(doc)
        try:
            os.startfile(pdf_path, "print")
        except Exception as ex:
            messagebox.showerror("Fehler", f"Drucken fehlgeschlagen: {ex}")

    # ===================== MENÜ-ACTIONEN =====================
    def _open_doc_overview(self):
        DocSearchDialog(self)

    def _open_customer_mgmt(self):
        CustomerDatabase(self)

    def _open_tool_mgmt(self):
        ToolDatabase(self)

    def _open_material_mgmt(self):
        MaterialDatabase(self)

    def _open_text_mgmt(self):
        TextDatabase(self)

    def _open_settings(self):
        SettingsDialog(self, master_mode=self._master_mode)

    def _backup_data(self):
        from tkinter import filedialog
        path = filedialog.asksaveasfilename(
            title="Datensicherung speichern",
            defaultextension=".db",
            filetypes=[("Datenbank", "*.db")],
            initialfile=f"FerdlWorks_Backup_{datetime.now().strftime('%Y%m%d')}.db")
        if path:
            if self.db.backup_to(path):
                messagebox.showinfo("Sicherung", "Datenbank gesichert.")

    def _restore_data(self):
        from tkinter import filedialog
        path = filedialog.askopenfilename(
            title="Datensicherung wiederherstellen",
            filetypes=[("Datenbank", "*.db")])
        if path:
            if messagebox.askyesno("Wiederherstellen", "Aktuelle Daten werden überschrieben. Fortfahren?"):
                if self.db.restore_from(path):
                    messagebox.showinfo("Wiederherstellung", "Datenbank wiederhergestellt.")
                    self._new_doc()

    def _cloud_gdrive(self):
        settings = self.db.settings_get_all()
        if not settings.get("gdrive_refresh_token"):
            if messagebox.askyesno("Google Drive", "Noch nicht autorisiert. Jetzt einrichten?"):
                success, msg = gdrive_authorize()
                if success:
                    messagebox.showinfo("Erfolg", msg)
                else:
                    messagebox.showerror("Fehler", msg)
            return
        success, msg = gdrive_backup(settings)
        if success:
            messagebox.showinfo("Google Drive", msg)
        else:
            messagebox.showerror("Fehler", msg)

    def _cloud_onedrive(self):
        settings = self.db.settings_get_all()
        if not settings.get("onedrive_refresh_token"):
            if messagebox.askyesno("OneDrive", "Noch nicht autorisiert. Jetzt einrichten?"):
                success, msg = onedrive_authorize()
                if success:
                    messagebox.showinfo("Erfolg", msg)
                else:
                    messagebox.showerror("Fehler", msg)
            return
        success, msg = onedrive_backup(settings)
        if success:
            messagebox.showinfo("OneDrive", msg)
        else:
            messagebox.showerror("Fehler", msg)

    def _open_log(self):
        log_path = get_log_path()
        try:
            os.startfile(log_path)
            self.logger.info(f"Logdatei geöffnet: {log_path}")
        except Exception as ex:
            self.logger.error(f"Konnte Logdatei nicht öffnen: {ex}")

    def _import_materials(self):
        from tkinter import filedialog, messagebox
        path = filedialog.askopenfilename(
            title="Excel-Datei ausw\u00e4hlen",
            filetypes=[("Excel-Dateien", "*.xlsx *.xls"), ("Alle Dateien", "*.*")]
        )
        if not path:
            return
        try:
            imported, skipped, errors = self.db.material_import_from_excel(path)
            msg = f"{imported} Materialien importiert."
            if skipped:
                msg += f"\n{skipped} leere Zeilen \u00fcbersprungen."
            if errors:
                msg += f"\n\nFehler:\n" + "\n".join(errors[:10])
                if len(errors) > 10:
                    msg += f"\n... und {len(errors) - 10} weitere"
            messagebox.showinfo("Import abgeschlossen", msg)
        except Exception as ex:
            messagebox.showerror("Import-Fehler", f"Fehler beim Import:\n{ex}")

    def _open_material_template(self):
        from tkinter import messagebox
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Materialien"
            ws.page_setup.orientation = "landscape"
            ws.merge_cells("A1:E1")
            c = ws["A1"]
            c.value = "Jede Zeile = ein Material. Einfach unterhalb der Beispiele eintragen."
            c.font = Font(italic=True, color="2F5496", size=10)
            c.alignment = Alignment(horizontal="left")
            headers = ["Name", "Beschreibung", "Preis", "Einheit", "Notiz"]
            widths = [30, 50, 15, 10, 40]
            hfont = Font(bold=True, color="FFFFFF", size=11)
            hfill = PatternFill("solid", fgColor="2F5496")
            halign = Alignment(horizontal="center", vertical="center")
            thin = Side(style="thin")
            border = Border(top=thin, left=thin, right=thin, bottom=thin)
            lfill = PatternFill("solid", fgColor="D6E4F0")
            for ci, (h, w) in enumerate(zip(headers, widths), 1):
                c = ws.cell(row=2, column=ci, value=h)
                c.font = hfont; c.fill = hfill; c.alignment = halign; c.border = border
                ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w
            examples = [
                ["Eiche Natur", "Massivholz Eiche, naturbelassen, 20mm", 89.50, "m\u00b2", "Innenbereich"],
                ["Buche Hell", "Buche Leimholz, gehobelt, 18mm", 72.00, "m\u00b2", ""],
                ["Schrauben 5x60", "Senkkopf, verzinkt, T20", 12.50, "Stk", "5er-Pack"],
            ]
            for ri, ex in enumerate(examples, 3):
                for ci, v in enumerate(ex, 1):
                    c = ws.cell(row=ri, column=ci, value=v)
                    c.border = border
                    if ri % 2 == 1:
                        c.fill = lfill
                    if ci == 3:
                        c.number_format = '#,##0.00'
            ws.freeze_panes = "A3"
            dst = os.path.join(os.environ["TEMP"], "materials_vorlage.xlsx")
            wb.save(dst)
            os.startfile(dst)
        except Exception as ex:
            messagebox.showerror("Fehler", f"Vorlage konnte nicht erstellt werden:\n{ex}")

    def _check_update(self):
        self.logger.info("Update-Prüfung gestartet")
        release, error = check_for_update()
        if error:
            messagebox.showerror("Update-Fehler", f"Konnte nicht nach Updates suchen:\n{error}")
            return
        if release is None:
            messagebox.showinfo("Aktuell", f"Sie haben die aktuellste Version v{VERSION}.")
            return
        tag = release.get("tag_name", "").lstrip("v")
        msg = f"Installierte Version: v{VERSION}\nNeue Version: v{tag}\n\nJetzt herunterladen und installieren?"
        if not messagebox.askyesno("Update verfügbar", msg):
            return
        dlg = ProgressDialog(self, "Update wird heruntergeladen...")
        dlg.show()
        path = download_installer(release, lambda p: dlg.set_progress(p))
        dlg.close()
        if path:
            app_path = reg_read("InstallPath", "")
            if app_path:
                app_exe = os.path.join(app_path, "FerdlWorks.exe")
            else:
                app_exe = sys.executable
            if install_and_restart(path, app_exe):
                self.logger.info(f"Update auf v{tag} gestartet, Neustart folgt")
                self.destroy()
            else:
                messagebox.showerror("Fehler", "Installation fehlgeschlagen.")
        else:
            messagebox.showerror("Fehler", "Download fehlgeschlagen.")

    def _uninstall(self):
        result = messagebox.askyesnocancel(
            "Deinstallation",
            "Möchten Sie FerdlWorks deinstallieren?\n\nJa = Einstellungen behalten\nNein = Alles löschen\nAbbrechen = Abbrechen")
        if result is None:
            return
        if not result:
            reg_delete_all()
        uninstall_path = reg_read("UninstallPath", "")
        if uninstall_path and os.path.exists(uninstall_path):
            subprocess.Popen([uninstall_path, "/SILENT"])
            self.destroy()
        else:
            messagebox.showerror("Fehler", "Kein Deinstallationsprogramm gefunden.")

    def _show_info(self):
        win = ctk.CTkToplevel(self)
        win.title("Info")
        win.geometry("350x200")
        win.resizable(False, False)
        win.transient(self)
        win.grab_set()
        ctk.CTkLabel(win, text=APP_NAME, font=("Segoe UI", 18, "bold"),
                     text_color=("#8b0000", "#8b0000")).pack(pady=(20, 5))
        ctk.CTkLabel(win, text=f"Version {VERSION}", font=("Segoe UI", 12)).pack()
        ctk.CTkLabel(win, text=COMPANY_NAME, font=("Segoe UI", 13),
                     text_color="#888888").pack(pady=(10, 0))
        ctk.CTkLabel(win, text="© 2026 Sonderegger Software", font=("Segoe UI", 13),
                     text_color="#666666").pack(pady=(5, 0))

    def _on_close(self):
        self.logger.info("Anwendung wird beendet")
        self.destroy()
        sys.exit(0)


# ===================== DOKUMENT-SUCHE =====================
class DocSearchDialog(ctk.CTkToplevel):
    def __init__(self, master, overdue_only=False):
        super().__init__(master)
        self.db = get_db()
        self.app = master
        self._overdue_only = overdue_only
        self.title("Rechnungen")
        self.geometry("800x500")
        self.transient(master)
        self.grab_set()
        self.after(50, lambda: set_window_icon(self, self.master))
        self._build_ui()
        self._load_data()

    def _build_ui(self):
        top = ctk.CTkFrame(self)
        top.pack(fill="x", padx=10, pady=10)
        ctk.CTkLabel(top, text="Suchen:").pack(side="left", padx=5)
        self.search_entry = ctk.CTkEntry(top, width=200, placeholder_text="Nr., Kunde...")
        self.search_entry.pack(side="left", padx=5)
        self.search_entry.bind("<KeyRelease>", lambda e: self._load_data())
        self.overdue_var = ctk.BooleanVar(value=self._overdue_only)
        ctk.CTkCheckBox(top, text="Überfällige Rechnungen", variable=self.overdue_var,
                        command=self._load_data).pack(side="left", padx=(15, 5))
        ctk.CTkButton(top, text="Öffnen", command=self._open_selected, width=80).pack(side="right", padx=5)
        ctk.CTkButton(top, text="Schließen", command=self.destroy, width=80).pack(side="right", padx=5)
        cols = ("doc", "kunde", "datum", "faellig", "betrag", "status")
        heads = {"doc": "Dokument", "kunde": "Kunde", "datum": "Datum",
                 "faellig": "Fällig", "betrag": "Betrag", "status": "Status"}
        widths = {"doc": 150, "kunde": 180, "datum": 85, "faellig": 85, "betrag": 90, "status": 70}
        self.tree = ttk.Treeview(self, columns=cols, show="headings")
        left_cols = ("doc", "kunde", "datum", "faellig")
        for c in cols:
            self.tree.heading(c, text=heads[c], anchor="w" if c in left_cols else "e")
            self.tree.column(c, width=widths[c], minwidth=50, anchor="w" if c in left_cols else "e")
        self.tree.column("status", width=70, anchor="center")
        self.tree.bind("<Double-1>", lambda e: self._open_selected())
        self.tree.bind("<Button-3>", self._tree_context_menu)
        vsb = ttk.Scrollbar(self, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        self.tree.pack(fill="both", expand=True, padx=10, pady=(0, 10))
        vsb.pack(side="right", fill="y")
        vsb.place(in_=self.tree, relx=1.0, rely=0, relheight=1.0, x=0)
        self._docs = []

    def _load_data(self):
        for row in self.tree.get_children():
            self.tree.delete(row)
        self._docs.clear()
        if self.overdue_var.get():
            self._docs = self.db.doc_get_overdue()
        else:
            query = self.search_entry.get()
            self._docs = self.db.doc_search(None, query)
        for doc in self._docs:
            company = doc.get("customer_name", "") or ""
            last = doc.get("customer_last_name", "") or ""
            first = doc.get("customer_first_name", "") or ""
            city = doc.get("customer_city", "") or ""
            name_part = company or f"{last} {first}".strip()
            cname = f"{name_part}, {city}" if city and name_part else name_part or "?"
            paid = doc.get("paid", "0") == "1"
            status = "\u2713" if paid else "\u2717"
            due = doc.get("due_date", "")
            self.tree.insert("", "end", iid=str(doc["id"]), values=(
                doc["doc_number"], cname, doc["date"],
                due, f"{doc.get('total_gross', 0):.2f}\u20ac".replace(".", ","), status
            ))

    def _tree_context_menu(self, event):
        iid = self.tree.identify_row(event.y)
        if not iid:
            return
        # select the row
        self.tree.selection_set(iid)
        doc_id = int(iid)
        doc = next((d for d in self._docs if d["id"] == doc_id), None)
        if not doc:
            return
        menu = tk.Menu(self, tearoff=False, font=("Segoe UI", 10))
        if doc["doc_type"] == "RG":
            is_paid = doc.get("paid", "0") == "1"
            menu.add_command(label="Als unbezahlt markieren" if is_paid else "Als bezahlt markieren",
                             command=lambda d=doc_id, p=not is_paid: self._toggle_paid(d, p))
            menu.add_separator()
        menu.add_command(label="Löschen", command=lambda d=doc_id: self._delete_doc(d))
        menu.tk_popup(event.x_root, event.y_root)
        menu.grab_release()

    def _toggle_paid(self, doc_id, paid):
        self.db.doc_set_paid(doc_id, paid)
        self._load_data()

    def _delete_doc(self, doc_id):
        doc = next((d for d in self._docs if d["id"] == doc_id), None)
        if not doc:
            return
        if not messagebox.askyesno("Löschen", f"{doc['doc_number']} wirklich löschen?\n\nDas Dokument und die zugehörige PDF-Datei werden endgültig gelöscht."):
            return
        self.db.doc_delete(doc_id)
        try:
            doc_type = doc["doc_type"]
            num = doc["doc_number"].replace(f"{doc_type}-", "").replace("-", "_")
            safe_date = doc.get("date", "").replace("-", "")
            pdf_name = f"{doc_type}_{num}_{safe_date}.pdf"
            pdf_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "data", "pdfs")
            pdf_path = os.path.join(pdf_dir, pdf_name)
            if os.path.exists(pdf_path):
                os.remove(pdf_path)
        except Exception:
            pass
        self._load_data()

    def _open_selected(self):
        sel = self.tree.selection()
        if not sel:
            return
        doc_id = int(sel[0])
        self.app._load_doc(doc_id)
        self.destroy()


class ProgressDialog:
    def __init__(self, master, title="Bitte warten..."):
        self.win = ctk.CTkToplevel(master)
        self.win.title(title)
        self.win.geometry("400x120")
        self.win.resizable(False, False)
        self.win.transient(master)
        self.win.grab_set()
        self.label = ctk.CTkLabel(self.win, text="Vorgang läuft...", font=("Segoe UI", 13))
        self.label.pack(pady=(15, 5))
        self.progress = ctk.CTkProgressBar(self.win, width=350, height=20,
                                           fg_color="#2a2a2a", progress_color="#8b0000")
        self.progress.pack(pady=10)
        self.progress.set(0)
        self.win.update()

    def show(self):
        self.win.deiconify()

    def set_progress(self, value):
        self.progress.set(value / 100)
        self.label.configure(text=f"{value:.0f}%")
        self.win.update()

    def close(self):
        try:
            self.win.destroy()
        except Exception:
            pass


def run_app():
    setup_logger()
    ctk.set_appearance_mode("dark")
    ctk.set_default_color_theme(THEME_PATH)
    root = ctk.CTk()
    root.withdraw()
    db = get_db()
    settings = db.settings_get_all()
    has_password = bool(settings.get("user_password", ""))
    if has_password:
        from lib.login_dialog import LoginDialog
        login = LoginDialog(root)
        root.wait_window(login)
        if not login.is_authenticated():
            root.destroy()
            return
        master_mode = login.is_master_mode()
    else:
        master_mode = False
    root.destroy()
    app = FerdlWorksApp(master_mode=master_mode)
    app.mainloop()


if __name__ == "__main__":
    run_app()
